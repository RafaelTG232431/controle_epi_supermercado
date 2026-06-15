from flask import Flask, render_template, request, redirect, url_for, session, send_file
from io import BytesIO
from database import db
from models import ControleAcesso, Colaborador, EPI, PedidoEPI, PedidoAlmoxarifado
from config import obter_banco
from openpyxl import Workbook, load_workbook
from datetime import timedelta
import os

app = Flask(__name__)

app.secret_key = "chave_secreta_supermercado_bh"

app.config["SQLALCHEMY_DATABASE_URI"] = obter_banco()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)

ADMIN_LOGIN = "LOJA522"
ADMIN_SENHA = "BH2026"

def atualizar_excel_pedidos():
    if not os.path.exists("relatorios"):
        os.makedirs("relatorios")

    arquivo = "relatorios/pedidos_atualizados.xlsx"

    pedidos = PedidoEPI.query.order_by(PedidoEPI.data.asc()).all()

    planilha = Workbook()

    aba_historico = planilha.active
    aba_historico.title = "Historico Pedidos"

    aba_historico.append([
        "Data do Pedido",
        "Nome",
        "Matrícula",
        "Tipo do Item",
        "Descrição",
        "Tamanho",
        "Quantidade",
        "Status",
        "Próxima Troca"
    ])

    pedidos_atuais = {}

    for pedido in pedidos:
        proxima_troca = pedido.data + timedelta(days=90)

        aba_historico.append([
            pedido.data.strftime("%d/%m/%Y %H:%M"),
            pedido.nome,
            pedido.matricula,
            pedido.tipo_item,
            pedido.descricao_item,
            pedido.tamanho,
            pedido.quantidade,
            pedido.status,
            proxima_troca.strftime("%d/%m/%Y")
        ])

        chave = f"{pedido.matricula}_{pedido.tipo_item}"

        pedidos_atuais[chave] = pedido

    aba_atual = planilha.create_sheet("Pedidos Atuais")

    aba_atual.append([
        "Data do Último Pedido",
        "Nome",
        "Matrícula",
        "Tipo do Item",
        "Descrição",
        "Tamanho",
        "Quantidade",
        "Status",
        "Próxima Troca"
    ])

    for pedido in pedidos_atuais.values():
        proxima_troca = pedido.data + timedelta(days=90)

        aba_atual.append([
            pedido.data.strftime("%d/%m/%Y %H:%M"),
            pedido.nome,
            pedido.matricula,
            pedido.tipo_item,
            pedido.descricao_item,
            pedido.tamanho,
            pedido.quantidade,
            pedido.status,
            proxima_troca.strftime("%d/%m/%Y")
        ])

    planilha.save(arquivo)

def iniciar_banco():
    with app.app_context():
        db.create_all()

        if ControleAcesso.query.first() is None:
            acesso = ControleAcesso(colaborador_liberado=False)
            db.session.add(acesso)

        if EPI.query.first() is None:
            itens_epi = [
                ("279715", "AVENTAL NAPA G UN", "Avental", "UN"),
                ("105038", "AVENTAL NAPA M UN", "Avental", "UN"),
                ("279713", "AVENTAL NAPA P UN", "Avental", "UN"),
                ("147039", "BOTA BICO ACO N33 PR", "Calçado", "PR"),
                ("65450", "BOTA BICO ACO N34 PR", "Calçado", "PR"),
                ("65240", "BOTA BICO ACO N35 PR", "Calçado", "PR"),
                ("65243", "BOTA BICO ACO N36 PR", "Calçado", "PR"),
                ("65241", "BOTA BICO ACO N37 PR", "Calçado", "PR"),
                ("65236", "BOTA BICO ACO N38 PR", "Calçado", "PR"),
                ("65239", "BOTA BICO ACO N39 PR", "Calçado", "PR"),
                ("65237", "BOTA BICO ACO N40 PR", "Calçado", "PR"),
                ("65407", "BOTA BICO ACO N41 PR", "Calçado", "PR"),
                ("65406", "BOTA BICO ACO N42 PR", "Calçado", "PR"),
                ("21912", "BOTA BICO ACO N43 PR", "Calçado", "PR"),
                ("65667", "BOTA BICO ACO N44 PR", "Calçado", "PR"),
                ("65404", "BOTA BICO ACO N45 PR", "Calçado", "PR"),
                ("147100", "BOTA BICO ACO N46 PR", "Calçado", "PR"),
                ("98629", "LUVA LATEX G PR", "Luva", "PR"),
                ("98628", "LUVA LATEX M PR", "Luva", "PR"),
                ("98621", "LUVA LATEX P PR", "Luva", "PR"),
                ("74202", "OCULOS SEGURANCA INCOLOR UN", "Óculos", "UN"),
                ("69652", "PROTETOR AURICULAR UN", "Protetor", "UN"),
                ("69918", "TOUCA LA NINJA UN", "Touca", "UN"),
            ]

            for codigo, descricao, categoria, unidade in itens_epi:
                epi = EPI(
                    codigo=codigo,
                    descricao=descricao,
                    categoria=categoria,
                    unidade=unidade
                )
                db.session.add(epi)

        if Colaborador.query.first() is None:
            colaboradores = [
                ("1001", "João Silva", "Açougue", "Açougueiro"),
                ("1002", "Maria Santos", "Padaria", "Atendente"),
                ("1003", "Carlos Lima", "Hortifruti", "Repositor"),
                ("1004", "Ana Costa", "Administrativo", "Auxiliar Administrativo"),
            ]

            for matricula, nome, setor, cargo in colaboradores:
                colaborador = Colaborador(
                    matricula=matricula,
                    nome=nome,
                    setor=setor,
                    cargo=cargo
                )
                db.session.add(colaborador)

        db.session.commit()

        print("BANCO DE DADOS CRIADO OU VERIFICADO COM SUCESSO.")


@app.route("/")
def login():
    return render_template("login.html")


@app.route("/login_admin", methods=["POST"])
def login_admin():
    usuario = request.form["usuario"]
    senha = request.form["senha"]

    if usuario == ADMIN_LOGIN and senha == ADMIN_SENHA:
        session["admin"] = True
        return redirect(url_for("admin"))

    return redirect(url_for("login"))


@app.route("/admin")
def admin():
    if not session.get("admin"):
        return redirect(url_for("login"))

    controle = ControleAcesso.query.first()

    total_colaboradores = Colaborador.query.count()
    total_epis = EPI.query.count()
    pedidos_epi = PedidoEPI.query.count()
    pedidos_almox = PedidoAlmoxarifado.query.count()
    pendentes = PedidoEPI.query.filter_by(status="Solicitado").count()

    return render_template(
        "admin.html",
        status=controle.colaborador_liberado,
        total_colaboradores=total_colaboradores,
        total_epis=total_epis,
        pedidos_epi=pedidos_epi,
        pedidos_almox=pedidos_almox,
        pendentes=pendentes
    )


@app.route("/liberar_colaborador")
def liberar_colaborador():
    if not session.get("admin"):
        return redirect(url_for("login"))

    controle = ControleAcesso.query.first()
    controle.colaborador_liberado = True
    db.session.commit()

    return redirect(url_for("admin"))


@app.route("/bloquear_colaborador")
def bloquear_colaborador():
    if not session.get("admin"):
        return redirect(url_for("login"))

    controle = ControleAcesso.query.first()
    controle.colaborador_liberado = False
    db.session.commit()

    return redirect(url_for("admin"))


@app.route("/acesso_colaborador")
def acesso_colaborador():
    controle = ControleAcesso.query.first()

    if not controle.colaborador_liberado:
        return render_template("bloqueado.html")

    return redirect(url_for("colaborador"))


@app.route("/colaborador")
def colaborador():
    controle = ControleAcesso.query.first()

    if not controle.colaborador_liberado:
        return render_template("bloqueado.html")

    return render_template("colaborador.html")


@app.route("/pedido_epi", methods=["POST"])
def pedido_epi():
    nome = request.form["nome"]
    matricula = request.form["matricula"]
    tipo_item = request.form["tipo_item"]
    quantidade = int(request.form["quantidade"])

    if tipo_item == "Outro":
        descricao_item = request.form["descricao_outro"]
    else:
        descricao_item = tipo_item

    tamanho = request.form["tamanho"]

    if tamanho == "Outro":
        tamanho = request.form["tamanho_outro"]

    pedido = PedidoEPI(
        nome=nome,
        matricula=matricula,
        tipo_item=tipo_item,
        descricao_item=descricao_item,
        tamanho=tamanho,
        quantidade=quantidade,
        status="Solicitado"
    )

    db.session.add(pedido)
    db.session.commit()

    atualizar_excel_pedidos()

    return redirect(url_for("colaborador"))


@app.route("/pedido_almoxarifado", methods=["POST"])
def pedido_almoxarifado():
    nome = request.form["nome"]
    matricula = request.form["matricula"]
    item_almoxarifado = request.form["item_almoxarifado"]
    quantidade = int(request.form["quantidade"])
    observacao = request.form["observacao"]

    if item_almoxarifado == "Outro":
        item = request.form["item_outro"]
    else:
        item = item_almoxarifado

    pedido = PedidoAlmoxarifado(
        matricula=matricula,
        solicitante=nome,
        setor="Não informado",
        item=item,
        quantidade=quantidade,
        observacao=observacao,
        status="Solicitado"
    )

    db.session.add(pedido)
    db.session.commit()

    atualizar_excel_pedidos()

    return redirect(url_for("colaborador"))

@app.route("/colaboradores")
def colaboradores():
    if not session.get("admin"):
        return redirect(url_for("login"))

    lista = Colaborador.query.order_by(Colaborador.nome).all()
    return render_template("colaboradores.html", colaboradores=lista)


@app.route("/cadastrar_colaborador", methods=["POST"])
def cadastrar_colaborador():
    if not session.get("admin"):
        return redirect(url_for("login"))

    novo = Colaborador(
        matricula=request.form["matricula"],
        nome=request.form["nome"],
        setor=request.form["setor"],
        cargo=request.form["cargo"]
    )

    db.session.add(novo)
    db.session.commit()

    return redirect(url_for("colaboradores"))


@app.route("/epis")
def epis():
    if not session.get("admin"):
        return redirect(url_for("login"))

    lista = EPI.query.order_by(EPI.descricao).all()
    return render_template("epis.html", epis=lista)


@app.route("/cadastrar_epi", methods=["POST"])
def cadastrar_epi():
    if not session.get("admin"):
        return redirect(url_for("login"))

    novo = EPI(
        codigo=request.form["codigo"],
        descricao=request.form["descricao"],
        categoria=request.form["categoria"],
        unidade=request.form["unidade"]
    )

    db.session.add(novo)
    db.session.commit()

    return redirect(url_for("epis"))


@app.route("/pedidos")
def pedidos():
    if not session.get("admin"):
        return redirect(url_for("login"))

    pedidos_epi = PedidoEPI.query.order_by(PedidoEPI.data.desc()).all()
    pedidos_almox = PedidoAlmoxarifado.query.order_by(PedidoAlmoxarifado.data.desc()).all()

    return render_template(
        "pedidos.html",
        pedidos_epi=pedidos_epi,
        pedidos_almox=pedidos_almox
    )


@app.route("/alterar_status_epi/<int:id>/<status>")
def alterar_status_epi(id, status):
    if not session.get("admin"):
        return redirect(url_for("login"))

    pedido = PedidoEPI.query.get(id)

    if pedido:
        pedido.status = status
        db.session.commit()

    return redirect(url_for("pedidos"))


@app.route("/exportar_excel")
def exportar_excel():
    if not session.get("admin"):
        return redirect(url_for("login"))

    planilha = Workbook()

    aba = planilha.active
    aba.title = "Pedidos EPI"

    aba.append([
        "Data",
        "Nome",
        "Matrícula",
        "Item",
        "Tamanho",
        "Quantidade",
        "Status"
    ])

    for pedido in PedidoEPI.query.order_by(PedidoEPI.data.desc()).all():
        aba.append([
            pedido.data.strftime("%d/%m/%Y %H:%M"),
            pedido.nome,
            pedido.matricula,
            pedido.descricao_item,
            pedido.tamanho,
            pedido.quantidade,
            pedido.status
        ])

    aba2 = planilha.create_sheet("Pedidos Almoxarifado")

    aba2.append([
        "Data",
        "Solicitante",
        "Matrícula",
        "Setor",
        "Item",
        "Quantidade",
        "Observação",
        "Status"
    ])

    for pedido in PedidoAlmoxarifado.query.order_by(PedidoAlmoxarifado.data.desc()).all():
        aba2.append([
            pedido.data.strftime("%d/%m/%Y %H:%M"),
            pedido.solicitante,
            pedido.matricula,
            pedido.setor,
            pedido.item,
            pedido.quantidade,
            pedido.observacao,
            pedido.status
        ])

    arquivo_memoria = BytesIO()
    planilha.save(arquivo_memoria)
    arquivo_memoria.seek(0)

    return send_file(
        arquivo_memoria,
        as_attachment=True,
        download_name="relatorio_pedidos_epi.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

with app.app_context():
    iniciar_banco()

@app.route("/sair")
def sair():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)